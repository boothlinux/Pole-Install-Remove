#!/usr/bin/python
# -*- coding: utf-8 -*-
import kivy  # Required to run Kivy such as the next line of code
kivy.require('1.9.1')  # used to alert user if this code is run on an earlier version of Kivy.
from kivy.app import App  # Imports the base App class required for Kivy Apps
from kivy.lang import Builder  # Imports the KV language builder that provides the layout of kivy screens
from kivy.uix.screenmanager import ScreenManager, Screen, NoTransition  # Imports the Kivy Screen manager and Kivys Screen class
from kivy.core.window import Window
import mainscreen
import poleremoval
import poleinstall
import streetlightremoval
import streetlightinstall
import transformerremoval
import transformerinstall
import rturemoval
import rtuinstall
import switchremoval
import switchinstall
import nettransremoval
import nettransinstall
import netproremoval
import netproinstall
import os
import dropbox
from openpyxl import load_workbook, drawing
import openpyxl
from kivy.properties import ObjectProperty
from kivy.uix.popup import Popup
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.image import Image
from kivy.uix.boxlayout import BoxLayout
import datepicker
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.scrollview import ScrollView

TOKEN = \
    'S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
dbx = dropbox.Dropbox(
    TOKEN,
    max_retries_on_error=10,
    max_retries_on_rate_limit=None,
    user_agent=None,
    session=None,
    headers=None,
    timeout=10000,
    )
Window.softinput_mode = 'below_target'
Builder.load_file('switchingscreen.kv')


class LoadDialog(FloatLayout):

    load = ObjectProperty(None)
    cancel = ObjectProperty(None)


class LoadDialog2(FloatLayout):

    load2 = ObjectProperty(None)
    cancel = ObjectProperty(None)


class PoleRemoval(Screen):

    def __init__(self, **kwargs):
        super(PoleRemoval, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleremoval.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleremoval.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        entera = self.ids.entera.text
        PMO = self.ids.PMO.text
        change = self.ids.change.text
        polenum = self.ids.polenum.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        locname = self.ids.locname.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        streetpos = self.ids.streetpos.text
        locstatus = self.ids.locstatus.text
        poleheightremoval = self.ids.poleheightremoval.text
        polematerial = self.ids.polematerial.text
        stub = self.ids.stub.text
        comments = self.ids.comments.text

        # Push to new excel file

        poleremoval.sheet.cell(row=2, column=2).value = '{}'.format(PMO)
        poleremoval.sheet.cell(row=2, column=6).value = \
            '{}'.format(change)
        poleremoval.sheet.cell(row=2, column=9).value = \
            '{}'.format('ENTERA')
        poleremoval.sheet.cell(row=3, column=2).value = \
            '{}'.format(polenum)
        poleremoval.sheet.cell(row=3, column=5).value = \
            '{}'.format(funcnum)
        poleremoval.sheet.cell(row=3, column=9).value = \
            '{}'.format(locnum)
        poleremoval.sheet.cell(row=4, column=2).value = \
            '{}'.format(locname)
        poleremoval.sheet.cell(row=5, column=2).value = \
            '{}'.format(streetnum)
        poleremoval.sheet.cell(row=5, column=5).value = \
            '{}'.format(streetname)
        poleremoval.sheet.cell(row=6, column=2).value = \
            '{}'.format(direction)
        poleremoval.sheet.cell(row=7, column=3).value = \
            '{}'.format(streetpos)
        poleremoval.sheet.cell(row=8, column=3).value = \
            '{}'.format(locstatus)
        poleremoval.sheet.cell(row=58, column=3).value = \
            '{}'.format(poleheightremoval)
        poleremoval.sheet.cell(row=58, column=6).value = \
            '{}'.format(polematerial)
        poleremoval.sheet.cell(row=58, column=10).value = \
            '{}'.format(stub)
        poleremoval.sheet.cell(row=60, column=1).value = \
            '{}'.format(comments)

        directory = os.path.abspath('/sdcard/UpDown')

        # directory = os.path.abspath("/home/cbooth/UpDown")

        if os.path.exists(directory):
            poleremoval.wb.save('{}/{} - Pole {} - {}.xlsx'.format(directory,
                                entera, polenum, change))
            file_from = '{}/{} - Pole {} - {}.xlsx'.format(directory,
                    entera, polenum, change)
            file_to = \
                '/SOFTWARE/{} - Pole {} - {} - {}.xlsx'.format(entera,
                    polenum, streetname, change)
        else:
            os.makedirs(directory)
            poleremoval.wb.save('{}/{} - Pole {} - {}.xlsx'.format(directory,
                                entera, polenum, change))
            file_from = '{}/{} - Pole {} - {}.xlsx'.format(directory,
                    entera, polenum, change)
            file_to = \
                '/SOFTWARE/{} - Pole {} - {} - {}.xlsx'.format(entera,
                    polenum, streetname, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)

        # file_from = ''

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''

        # self.ids.direction.text = ''

        self.ids.streetpos.text = ''
        self.ids.locstatus.text = ''
        self.ids.poleheightremoval.text = ''
        self.ids.polematerial.text = ''
        self.ids.stub.text = ''
        self.ids.comments.text = ''

    loadfile = ObjectProperty(None)
    savefile = ObjectProperty(None)
    text_input = ObjectProperty(None)


class PoleInstall(Screen):  # Creates a SecondScreen widget from the above kv language data string.

    def __init__(self, **kwargs):
        super(PoleInstall, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleinstall.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleinstall.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        polenum = self.ids.polenum.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        locname = self.ids.locname.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        streetpos = self.ids.streetpos.text
        locstatus = self.ids.locstatus.text
        poleowner = self.ids.poleowner.text
        poleheightinstall = self.ids.poleheightinstall.text
        stocknum = self.ids.stocknum.text
        materialinstall = self.ids.materialinstall.text
        polesetting = self.ids.polesetting.text
        poleframing = self.ids.poleframing.text
        surface = self.ids.surface.text
        jointuse = self.ids.jointuse.text
        traffic = self.ids.traffic.text
        topext = self.ids.topext.text
        feeder = self.ids.feeder.text
        streetlight = self.ids.streetlight.text
        riser = self.ids.riser.text
        suspension = self.ids.suspension.text
        signs = self.ids.signs.text
        guy = self.ids.guy.text

        # Push to new excel file

        poleinstall.sheet.cell(row=2, column=2).value = '{}'.format(PMO)
        poleinstall.sheet.cell(row=2, column=6).value = \
            '{}'.format(change)
        poleinstall.sheet.cell(row=2, column=9).value = \
            '{}'.format('ENTERA')
        poleinstall.sheet.cell(row=3, column=2).value = \
            '{}'.format(polenum)
        poleinstall.sheet.cell(row=3, column=5).value = \
            '{}'.format(funcnum)
        poleinstall.sheet.cell(row=3, column=9).value = \
            '{}'.format(locnum)
        poleinstall.sheet.cell(row=4, column=2).value = \
            '{}'.format(locname)
        poleinstall.sheet.cell(row=5, column=2).value = \
            '{}'.format(streetnum)
        poleinstall.sheet.cell(row=5, column=5).value = \
            '{}'.format(streetname)
        poleinstall.sheet.cell(row=6, column=2).value = \
            '{}'.format(direction)
        poleinstall.sheet.cell(row=7, column=3).value = \
            '{}'.format(streetpos)
        poleinstall.sheet.cell(row=8, column=3).value = \
            '{}'.format(locstatus)
        poleinstall.sheet.cell(row=10, column=3).value = \
            '{}'.format(poleowner)
        poleinstall.sheet.cell(row=11, column=3).value = \
            '{}'.format(poleheightinstall)
        poleinstall.sheet.cell(row=11, column=7).value = \
            '{}'.format(stocknum)
        poleinstall.sheet.cell(row=12, column=3).value = \
            '{}'.format(materialinstall)
        poleinstall.sheet.cell(row=13, column=3).value = \
            '{}'.format(polesetting)
        poleinstall.sheet.cell(row=14, column=3).value = \
            '{}'.format(poleframing)
        poleinstall.sheet.cell(row=15, column=3).value = \
            '{}'.format(surface)
        poleinstall.sheet.cell(row=15, column=7).value = \
            '{}'.format(jointuse)
        poleinstall.sheet.cell(row=15, column=9).value = \
            '{}'.format(traffic)
        poleinstall.sheet.cell(row=16, column=3).value = \
            '{}'.format(topext)
        poleinstall.sheet.cell(row=16, column=7).value = \
            '{}'.format(feeder)
        poleinstall.sheet.cell(row=16, column=9).value = \
            '{}'.format(streetlight)
        poleinstall.sheet.cell(row=17, column=3).value = \
            '{}'.format(riser)
        poleinstall.sheet.cell(row=17, column=7).value = \
            '{}'.format(suspension)
        poleinstall.sheet.cell(row=17, column=9).value = \
            '{}'.format(signs)
        poleinstall.sheet.cell(row=18, column=3).value = \
            '{}'.format(guy)

        # SET DIRECTORY

        directory = os.path.abspath('/sdcard/UpDown')  # FOR ANDROID VERSION

        # directory = os.path.abspath("/home/cbooth/UpDown") # FOR LINUX VERSION

        if os.path.exists(directory):
            poleinstall.wb.save('{}/Pole {} - {}.xlsx'.format(directory,
                                polenum, change))
            file_from = '{}/Pole {} - {}.xlsx'.format(directory,
                    polenum, change)
            file_to = \
                '/SOFTWARE/Pole {} - {} - {}.xlsx'.format(polenum,
                    streetname, change)
        else:
            os.makedirs(directory)
            poleinstall.wb.save('{}/Pole {} - {}.xlsx'.format(directory,
                                polenum, change))
            file_from = '{}/Pole {} - {} - {}.xlsx'.format(directory,
                    polenum, streetname, change)
            file_to = \
                '/SOFTWARE/Pole {} - {} - {}.xlsx'.format(polenum,
                    streetname, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.direction.text = ''
        self.ids.streetpos.text = ''
        self.ids.locstatus.text = ''


class StreetlightRemoval(Screen):

    def __init__(self, **kwargs):
        super(StreetlightRemoval, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        streetlightremoval.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        streetlightremoval.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        pass

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        polenum = self.ids.polenum.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        region = self.ids.region.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        locstatus = self.ids.locstatus.text
        owner = self.ids.owner.text
        addloc = self.ids.addloc.text
        fixremove = self.ids.fixremove.text
        bracketremove = self.ids.bracketremove.text


        # Push to new excel file

        streetlightremoval.sheet.cell(row=2, column=2).value = '{}'.format(PMO)
        streetlightremoval.sheet.cell(row=2, column=4).value = \
            '{}'.format(change)
        streetlightremoval.sheet.cell(row=2, column=7).value = \
            '{}'.format('ENTERA')
        streetlightremoval.sheet.cell(row=3, column=6).value = \
            '{}'.format(polenum)
        streetlightremoval.sheet.cell(row=3, column=2).value = \
            '{}'.format(funcnum)
        streetlightremoval.sheet.cell(row=5, column=2).value = \
            '{}'.format(locnum)
        streetlightremoval.sheet.cell(row=5, column=5).value = \
            '{}'.format(region)
        streetlightremoval.sheet.cell(row=4, column=2).value = \
            '{}'.format(streetnum)
        streetlightremoval.sheet.cell(row=4, column=4).value = \
            '{}'.format(streetname)
        streetlightremoval.sheet.cell(row=6, column=2).value = \
            '{}'.format(addloc)
        streetlightremoval.sheet.cell(row=8, column=2).value = \
            '{}'.format(owner)
        streetlightremoval.sheet.cell(row=8, column=3).value = \
            '{}'.format(locstatus)
        streetlightremoval.sheet.cell(row=20, column=1).value = \
            '{}'.format(fixremove)
        streetlightremoval.sheet.cell(row=20, column=5).value = \
            '{}'.format(bracketremove)

        # SET DIRECTORY
        directory = os.path.abspath("/sdcard/UpDown") #FOR ANDROID VERSION

        #directory = os.path.abspath('/home/cbooth/UpDown')  # FOR LINUX VERSION
        
        if os.path.exists(directory):
            streetlightremoval.wb.save('{}/Streetlight {} - {}.xlsx'.format(directory,
                    polenum, change))
            file_from = '{}/Streetlight {} - {}.xlsx'.format(directory,
                    polenum, change)
            file_to = \
                '/SOFTWARE/Streetlight {} - {} - {}.xlsx'.format(polenum,
                    streetname, change)
        else:
            os.makedirs(directory)
            streetlightremoval.wb.save('{}/Streetlight {} - {}.xlsx'.format(directory,
                    polenum, change))
            file_from = '{}/Streetlight {} - {}.xlsx'.format(directory,
                    polenum, change)
            file_to = \
                '/SOFTWARE/Streetlight {} - {} - {}.xlsx'.format(polenum,
                    streetname, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.region.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.locstatus.text = ''
        self.ids.owner.text = ''
        self.ids.addloc.text = ''
        self.ids.fixremove.text = ''
        self.ids.bracketremove.text = ''


class StreetlightInstall(Screen):

    def __init__(self, **kwargs):
        super(StreetlightInstall, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        streetlightinstall.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        streetlightinstall.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        fixremove = self.ids.fixremove.text
        bracketremove = self.ids.bracketremove.text
        owner = self.ids.owner.text
        locstatus = self.ids.locstatus.text
        region = self.ids.region.text
        
    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        polenum = self.ids.polenum.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        region = self.ids.region.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        locstatus = self.ids.locstatus.text
        owner = self.ids.owner.text
        addloc = self.ids.addloc.text
        fixremove = self.ids.fixremove.text
        bracketremove = self.ids.bracketremove.text

        # Push to new excel file

        streetlightremoval.sheet.cell(row=2, column=2).value = '{}'.format(PMO)
        
        streetlightremoval.sheet.cell(row=2, column=4).value = \
            '{}'.format(change)
        streetlightremoval.sheet.cell(row=2, column=7).value = \
            '{}'.format('ENTERA')
        streetlightremoval.sheet.cell(row=3, column=6).value = \
            '{}'.format(polenum)
        streetlightremoval.sheet.cell(row=3, column=2).value = \
            '{}'.format(funcnum)
        streetlightremoval.sheet.cell(row=5, column=2).value = \
            '{}'.format(locnum)
        streetlightremoval.sheet.cell(row=5, column=5).value = \
            '{}'.format(region)
        streetlightremoval.sheet.cell(row=4, column=2).value = \
            '{}'.format(streetnum)
        streetlightremoval.sheet.cell(row=4, column=4).value = \
            '{}'.format(streetname)
        streetlightremoval.sheet.cell(row=6, column=2).value = \
            '{}'.format(addloc)
        streetlightremoval.sheet.cell(row=8, column=2).value = \
            '{}'.format(owner)
        streetlightremoval.sheet.cell(row=8, column=3).value = \
            '{}'.format(locstatus)

        # SET DIRECTORY

        directory = os.path.abspath('/sdcard/UpDown')  # FOR ANDROID VERSION

        # directory = os.path.abspath("/home/cbooth/UpDown") # FOR LINUX VERSION

        if os.path.exists(directory):
            streetlightremove.wb.save('{}/Streetlight {} - {}.xlsx'.format(directory,
                    polenum, change))
            file_from = '{}/Streetlight {} - {}.xlsx'.format(directory,
                    polenum, change)
            file_to = \
                '/SOFTWARE/Streetlight {} - {} - {}.xlsx'.format(polenum,
                    streetname, change)
        else:
            os.makedirs(directory)
            streetlightremove.wb.save('{}/Streetlight {} - {}.xlsx'.format(directory,
                    polenum, change))
            file_from = '{}/Streetlight {} - {}.xlsx'.format(directory,
                    polenum, change)
            file_to = \
                '/SOFTWARE/Streetlight {} - {} - {}.xlsx'.format(polenum,
                    streetname, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.region.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.locstatus.text = ''
        self.ids.owner.text = ''
        self.ids.addloc.text = ''
        self.ids.fixremove.text = ''
        self.ids.bracketremove.text = ''


class TransformerRemoval(Screen):

    def __init__(self, **kwargs):
        super(TransformerRemoval, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        transformerremoval.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        transformerremoval.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        polenum = self.ids.polenum.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        locname = self.ids.locname.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        streetpos = self.ids.streetpos.text
        locstatus = self.ids.locstatus.text
        poleowner = self.ids.poleowner.text
        poleheightinstall = self.ids.poleheightinstall.text
        stocknum = self.ids.stocknum.text
        materialinstall = self.ids.materialinstall.text
        polesetting = self.ids.polesetting.text
        poleframing = self.ids.poleframing.text
        surface = self.ids.surface.text
        jointuse = self.ids.jointuse.text
        traffic = self.ids.traffic.text
        topext = self.ids.topext.text
        feeder = self.ids.feeder.text
        streetlight = self.ids.streetlight.text
        riser = self.ids.riser.text
        suspension = self.ids.suspension.text
        signs = self.ids.signs.text
        guy = self.ids.guy.text

        # Push to new excel file

        transformerremoval.sheet.cell(row=2, column=5).value = \
            '{}'.format(PMO)
        transformerremoval.sheet.cell(row=2, column=14).value = \
            '{}'.format(change)
        transformerremoval.sheet.cell(row=2, column=25).value = \
            '{}'.format('ENTERA')
        transformerremoval.sheet.cell(row=3, column=5).value = \
            '{}'.format(funcnum)
        transformerremoval.sheet.cell(row=3, column=21).value = \
            '{}'.format(locnum)
        transformerremoval.sheet.cell(row=3, column=33).value = \
            '{}'.format(owner)
        transformerremoval.sheet.cell(row=4, column=4).value = \
            '{}'.format(polenum)
        transformerremoval.sheet.cell(row=4, column=12).value = \
            '{}'.format(feeder)
        transformerremoval.sheet.cell(row=4, column=21).value = \
            '{}'.format(region)
        transformerremoval.sheet.cell(row=5, column=4).value = \
            '{}'.format(streetnum)
        transformerremoval.sheet.cell(row=5, column=11).value = \
            '{}'.format(streetname)
        transformerremoval.sheet.cell(row=6, column=4).value = \
            '{}'.format(direction)
        transformerremoval.sheet.cell(row=7, column=6).value = \
            '{}'.format(streetpos)
        transformerremoval.sheet.cell(row=8, column=6).value = \
            '{}'.format(loctype)
        transformerremoval.sheet.cell(row=8, column=35).value = \
            '{}'.format(tempsvc)
        transformerremoval.sheet.cell(row=9, column=6).value = \
            '{}'.format(locstatus)
        transformerremoval.sheet.cell(row=10, column=16).value = \
            '{}'.format(ehstracking)           
        transformerremoval.sheet.cell(row=31, column=3).value = \
            '{}'.format(equip1)
        transformerremoval.sheet.cell(row=31, column=12).value = \
            '{}'.format(serial1)
        transformerremoval.sheet.cell(row=31, column=20).value = \
            '{}'.format(manu1)
        transformerremoval.sheet.cell(row=31, column=26).value = \
            '{}'.format(stock1)
        transformerremoval.sheet.cell(row=31, column=32).value = \
            '{}'.format(tsm1)
        transformerremoval.sheet.cell(row=32, column=3).value = \
            '{}'.format(equip2)
        transformerremoval.sheet.cell(row=32, column=12).value = \
            '{}'.format(serial2)
        transformerremoval.sheet.cell(row=32, column=20).value = \
            '{}'.format(manu2)
        transformerremoval.sheet.cell(row=32, column=26).value = \
            '{}'.format(stock2)
        transformerremoval.sheet.cell(row=32, column=32).value = \
            '{}'.format(tsm2)
        transformerremoval.sheet.cell(row=33, column=3).value = \
            '{}'.format(equip3)
        transformerremoval.sheet.cell(row=33, column=12).value = \
            '{}'.format(serial3)
        transformerremoval.sheet.cell(row=33, column=20).value = \
            '{}'.format(manu3)
        transformerremoval.sheet.cell(row=33, column=26).value = \
            '{}'.format(stock3)
        transformerremoval.sheet.cell(row=33, column=32).value = \
            '{}'.format(tsm3)
        transformerremoval.sheet.cell(row=34, column=10).value = \
            '{}'.format(status)
        transformerremoval.sheet.cell(row=35, column=10).value = \
            '{}'.format(equipreturn)

        # SET DIRECTORY

        directory = os.path.abspath('/sdcard/UpDown')  # FOR ANDROID VERSION

        # directory = os.path.abspath("/home/cbooth/UpDown") # FOR LINUX VERSION

        if os.path.exists(directory):
            streetlightremove.wb.save('{}/Transformer {} - {}.xlsx'.format(directory,
                    polenum, change))
            file_from = '{}/Transformer {} - {}.xlsx'.format(directory,
                    polenum, change)
            file_to = \
                '/SOFTWARE/Transformer {} - {} - {}.xlsx'.format(polenum,
                    streetname, change)
        else:
            os.makedirs(directory)
            streetlightremove.wb.save('{}/Transformer {} - {}.xlsx'.format(directory,
                    polenum, change))
            file_from = '{}/Transformer {} - {}.xlsx'.format(directory,
                    polenum, change)
            file_to = \
                '/SOFTWARE/Transformer {} - {} - {}.xlsx'.format(polenum,
                    streetname, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.direction.text = ''
        self.ids.streetpos.text = ''
        self.ids.locstatus.text = ''


class TransformerInstall(Screen):

    def __init__(self, **kwargs):
        super(TransformerInstall, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        transformerinstall.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        transformerinstall.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        polenum = self.ids.polenum.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        locname = self.ids.locname.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        streetpos = self.ids.streetpos.text
        locstatus = self.ids.locstatus.text
        poleowner = self.ids.poleowner.text
        poleheightinstall = self.ids.poleheightinstall.text
        stocknum = self.ids.stocknum.text
        materialinstall = self.ids.materialinstall.text
        polesetting = self.ids.polesetting.text
        poleframing = self.ids.poleframing.text
        surface = self.ids.surface.text
        jointuse = self.ids.jointuse.text
        traffic = self.ids.traffic.text
        topext = self.ids.topext.text
        feeder = self.ids.feeder.text
        streetlight = self.ids.streetlight.text
        riser = self.ids.riser.text
        suspension = self.ids.suspension.text
        signs = self.ids.signs.text
        guy = self.ids.guy.text

        # Push to new excel file

        transformerinstall.sheet.cell(row=2, column=5).value = \
            '{}'.format(PMO)
        transformerinstall.sheet.cell(row=2, column=14).value = \
            '{}'.format(change)
        transformerinstall.sheet.cell(row=2, column=25).value = \
            '{}'.format('ENTERA')
        transformerinstall.sheet.cell(row=3, column=5).value = \
            '{}'.format(funcnum)
        transformerinstall.sheet.cell(row=3, column=21).value = \
            '{}'.format(locnum)
        transformerinstall.sheet.cell(row=3, column=33).value = \
            '{}'.format(owner)
        transformerinstall.sheet.cell(row=4, column=4).value = \
            '{}'.format(polenum)
        transformerinstall.sheet.cell(row=4, column=12).value = \
            '{}'.format(feeder)
        transformerinstall.sheet.cell(row=4, column=21).value = \
            '{}'.format(region)
        transformerinstall.sheet.cell(row=5, column=4).value = \
            '{}'.format(streetnum)
        transformerinstall.sheet.cell(row=5, column=11).value = \
            '{}'.format(streetname)
        transformerinstall.sheet.cell(row=6, column=4).value = \
            '{}'.format(direction)
        transformerinstall.sheet.cell(row=7, column=6).value = \
            '{}'.format(streetpos)
        transformerinstall.sheet.cell(row=8, column=6).value = \
            '{}'.format(loctype)
        transformerinstall.sheet.cell(row=8, column=35).value = \
            '{}'.format(tempsvc)
        transformerinstall.sheet.cell(row=9, column=6).value = \
            '{}'.format(locstatus)
        transformerinstall.sheet.cell(row=10, column=16).value = \
            '{}'.format(ehstracking)
        transformerinstall.sheet.cell(row=12, column=6).value = \
            '{}'.format(equip1)
        transformerinstall.sheet.cell(row=13, column=6).value = \
            '{}'.format(serial1)
        transformerinstall.sheet.cell(row=14, column=6).value = \
            '{}'.format(manu1)
        transformerinstall.sheet.cell(row=14, column=15).value = \
            '{}'.format(stock1)
        transformerinstall.sheet.cell(row=15, column=4).value = \
            '{}'.format(kva1)
        transformerinstall.sheet.cell(row=15, column=9).value = \
            '{}'.format(imp1)
        transformerinstall.sheet.cell(row=15, column=14).value = \
            '{}'.format(phase1)
        transformerinstall.sheet.cell(row=16, column=7).value = \
            '{}'.format(tsm1)
        transformerinstall.sheet.cell(row=13, column=21).value = \
            '{}'.format(privolt1)
        transformerinstall.sheet.cell(row=13, column=25).value = \
            '{}'.format(priconfig1)
        transformerinstall.sheet.cell(row=13, column=28).value = \
            '{}'.format(secvolt1)
        transformerinstall.sheet.cell(row=13, column=32).value = \
            '{}'.format(secconfig1)
        transformerinstall.sheet.cell(row=13, column=25).value = \
            '{}'.format(tapset1)
        transformerinstall.sheet.cell(row=17, column=6).value = \
            '{}'.format(equip2)
        transformerinstall.sheet.cell(row=18, column=6).value = \
            '{}'.format(serial2)
        transformerinstall.sheet.cell(row=19, column=6).value = \
            '{}'.format(manu2)
        transformerinstall.sheet.cell(row=19, column=15).value = \
            '{}'.format(stock2)
        transformerinstall.sheet.cell(row=20, column=4).value = \
            '{}'.format(kva2)
        transformerinstall.sheet.cell(row=20, column=9).value = \
            '{}'.format(imp2)
        transformerinstall.sheet.cell(row=20, column=14).value = \
            '{}'.format(phase2)
        transformerinstall.sheet.cell(row=21, column=7).value = \
            '{}'.format(tsm2)
        transformerinstall.sheet.cell(row=18, column=2).value = \
            '{}'.format(privolt2)
        transformerinstall.sheet.cell(row=18, column=25).value = \
            '{}'.format(priconfig2)
        transformerinstall.sheet.cell(row=18, column=28).value = \
            '{}'.format(secvolt2)
        transformerinstall.sheet.cell(row=18, column=32).value = \
            '{}'.format(secconfig2)
        transformerinstall.sheet.cell(row=18, column=35).value = \
            '{}'.format(tapset2)
        transformerinstall.sheet.cell(row=22, column=6).value = \
            '{}'.format(equip3)
        transformerinstall.sheet.cell(row=23, column=6).value = \
            '{}'.format(serial3)
        transformerinstall.sheet.cell(row=24, column=6).value = \
            '{}'.format(manu3)
        transformerinstall.sheet.cell(row=24, column=15).value = \
            '{}'.format(stock3)
        transformerinstall.sheet.cell(row=25, column=4).value = \
            '{}'.format(kva3)
        transformerinstall.sheet.cell(row=25, column=9).value = \
            '{}'.format(imp3)
        transformerinstall.sheet.cell(row=25, column=14).value = \
            '{}'.format(phase3)
        transformerinstall.sheet.cell(row=26, column=7).value = \
            '{}'.format(tsm3)
        transformerinstall.sheet.cell(row=23, column=2).value = \
            '{}'.format(privolt3)
        transformerinstall.sheet.cell(row=23, column=25).value = \
            '{}'.format(priconfig3)
        transformerinstall.sheet.cell(row=23, column=28).value = \
            '{}'.format(secvolt3)
        transformerinstall.sheet.cell(row=23, column=32).value = \
            '{}'.format(secconfig3)
        transformerinstall.sheet.cell(row=23, column=35).value = \
            '{}'.format(tapset3)
        transformerinstall.sheet.cell(row=27, column=101).value = \
            '{}'.format(metertx1)
        transformerinstall.sheet.cell(row=27, column=20).value = \
            '{}'.format(metertx2)
        transformerinstall.sheet.cell(row=20, column=30).value = \
            '{}'.format(metertx3)
        transformerinstall.sheet.cell(row=28, column=7).value = \
            '{}'.format(testpoint)
        transformerinstall.sheet.cell(row=28, column=22).value = \
            '{}'.format(status)

        # SET DIRECTORY

        directory = os.path.abspath('/sdcard/UpDown')  # FOR ANDROID VERSION

        # directory = os.path.abspath("/home/cbooth/UpDown") # FOR LINUX VERSION

        if os.path.exists(directory):
            streetlightremove.wb.save('{}/Transformer {} - {}.xlsx'.format(directory,
                    polenum, change))
            file_from = '{}/Transformer {} - {}.xlsx'.format(directory,
                    polenum, change)
            file_to = \
                '/SOFTWARE/Transformer {} - {} - {}.xlsx'.format(polenum,
                    feeder, change)
        else:
            os.makedirs(directory)
            streetlightremove.wb.save('{}/Transformer {} - {}.xlsx'.format(directory,
                    polenum, change))
            file_from = '{}/Transformer {} - {}.xlsx'.format(directory,
                    polenum, change)
            file_to = \
                '/SOFTWARE/Transformer {} - {} - {}.xlsx'.format(polenum,
                    feeder, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.direction.text = ''
        self.ids.streetpos.text = ''
        self.ids.locstatus.text = ''


class RTURemoval(Screen):

    def __init__(self, **kwargs):
        super(RTURemoval, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        rturemoval.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        rturemoval.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        owner = self.ids.owner.text
        polenum = self.ids.polenum.text
        region = self.ids.region.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        loctype = self.ids.loctype.text
        locstatus = self.ids.locstatus.text

        # Push to new excel file

        rturemoval.sheet.cell(row=2, column=6).value = '{}'.format(PMO)
        rturemoval.sheet.cell(row=2, column=15).value = \
            '{}'.format(change)
        rturemoval.sheet.cell(row=2, column=26).value = \
            '{}'.format('ENTERA')
        rturemoval.sheet.cell(row=3, column=6).value = \
            '{}'.format(funcnum)
        rturemoval.sheet.cell(row=3, column=22).value = \
            '{}'.format(locnum)
        rturemoval.sheet.cell(row=3, column=34).value = \
            '{}'.format(owner)
        rturemoval.sheet.cell(row=4, column=5).value = \
            '{}'.format(polenum)
        rturemoval.sheet.cell(row=5, column=13).value = \
            '{}'.format(region)
        rturemoval.sheet.cell(row=5, column=5).value = \
            '{}'.format(streetnum)
        rturemoval.sheet.cell(row=5, column=10).value = \
            '{}'.format(streetname)
        rturemoval.sheet.cell(row=5, column=28).value = \
            '{}'.format(direction)
        rturemoval.sheet.cell(row=6, column=7).value = \
            '{}'.format(loctype)
        rturemoval.sheet.cell(row=7, column=7).value = \
            '{}'.format(locstatus)
        rturemoval.sheet.cell(row=21, column=1).value = \
            '{}'.format(stock)
        rturemoval.sheet.cell(row=21, column=8).value = \
            '{}'.format(processor)
        rturemoval.sheet.cell(row=21, column=13).value = \
            '{}'.format(serialnum)
        rturemoval.sheet.cell(row=21, column=19).value = \
            '{}'.format(chargermake)
        rturemoval.sheet.cell(row=21, column=24).value = \
            '{}'.format(chargermodel)
        rturemoval.sheet.cell(row=21, column=29).value = \
            '{}'.format(battmake)
        rturemoval.sheet.cell(row=21, column=34).value = \
            '{}'.format(battmodel)
        rturemoval.sheet.cell(row=22, column=11).value = \
            '{}'.format(status)
        rturemoval.sheet.cell(row=23, column=6).value = \
            '{}'.format(equipreturn)
        rturemoval.sheet.cell(row=25, column=1).value = \
            '{}'.format(comments)

        # SET DIRECTORY

        directory = os.path.abspath('/sdcard/UpDown')  # FOR ANDROID VERSION

        # directory = os.path.abspath("/home/cbooth/UpDown") # FOR LINUX VERSION

        if os.path.exists(directory):
            rturemoval.wb.save('{}/RTU {} - {}.xlsx'.format(directory,
                               polenum, change))
            file_from = '{}/RTU {} - {}.xlsx'.format(directory,
                    polenum, change)
            file_to = '/SOFTWARE/RTU {} - {} - {}.xlsx'.format(polenum,
                    streetname, change)
        else:
            os.makedirs(directory)
            rturemoval.wb.save('{}/RTU {} - {}.xlsx'.format(directory,
                               polenum, change))
            file_from = '{}/RTU {} - {}.xlsx'.format(directory,
                    polenum, change)
            file_to = '/SOFTWARE/RTU {} - {} - {}.xlsx'.format(polenum,
                    streetname, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.direction.text = ''
        self.ids.streetpos.text = ''
        self.ids.locstatus.text = ''


class RTUInstall(Screen):

    def __init__(self, **kwargs):
        super(RTUInstall, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        rtuinstall.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        rtuinstall.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        owner = self.ids.owner.text
        polenum = self.ids.polenum.text
        region = self.ids.region.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        loctype = self.ids.loctype.text
        locstatus = self.ids.locstatus.text
        enclosure = self.ids.enclosure.text
        equipnum = self.ids.equipnum.text
        serialnum = self.ids.serialnum.text
        manufacturer = self.ids.manufacturer.text
        acsource = self.ids.acsource.text
        stock = self.ids.stock.text
        communication = self.ids.communication.text
        rtuvolt = self.ids.rtuvolt.text
        processor = self.ids.processor.text
        chargermake = self.ids.chargermake.text
        chargermodel = self.ids.chargermodel.text
        chargervolt = self.ids.chargervolt.text
        chargeramp = self.ids.chargeramp.text
        battmake = self.ids.battmake.text
        battmodel = self.ids.battmodel.text
        battnum = self.ids.battnum.text
        battvolt = self.ids.battvolt.text
        batttype = self.ids.batttype.text
        battamp = self.ids.battamp.text
        comments = self.ids.comments.text

        # Push to new excel file

        switchremoval.sheet.cell(row=2, column=6).value = \
            '{}'.format(PMO)
        switchremoval.sheet.cell(row=2, column=15).value = \
            '{}'.format(change)
        switchremoval.sheet.cell(row=2, column=26).value = \
            '{}'.format('ENTERA')
        switchremoval.sheet.cell(row=3, column=6).value = \
            '{}'.format(funcnum)
        switchremoval.sheet.cell(row=3, column=22).value = \
            '{}'.format(locnum)
        switchremoval.sheet.cell(row=3, column=34).value = \
            '{}'.format(owner)
        switchremoval.sheet.cell(row=4, column=5).value = \
            '{}'.format(polenum)
        switchremoval.sheet.cell(row=5, column=13).value = \
            '{}'.format(region)
        switchremoval.sheet.cell(row=5, column=5).value = \
            '{}'.format(streetnum)
        switchremoval.sheet.cell(row=5, column=10).value = \
            '{}'.format(streetname)
        switchremoval.sheet.cell(row=5, column=28).value = \
            '{}'.format(direction)
        switchremoval.sheet.cell(row=6, column=7).value = \
            '{}'.format(loctype)
        switchremoval.sheet.cell(row=7, column=7).value = \
            '{}'.format(locstatus)
        switchremoval.sheet.cell(row=9, column=6).value = \
            '{}'.format(enclosure)
        switchremoval.sheet.cell(row=10, column=6).value = \
            '{}'.format(equipnum)
        switchremoval.sheet.cell(row=11, column=6).value = \
            '{}'.format(serialnum)
        switchremoval.sheet.cell(row=12, column=6).value = \
            '{}'.format(manufacturer)
        switchremoval.sheet.cell(row=13, column=6).value = \
            '{}'.format(acsource)
        switchremoval.sheet.cell(row=12, column=16).value = \
            '{}'.format(stock)
        switchremoval.sheet.cell(row=10, column=22).value = \
            '{}'.format(communication)
        switchremoval.sheet.cell(row=10, column=27).value = \
            '{}'.format(rtuvolt)
        switchremoval.sheet.cell(row=10, column=30).value = \
            '{}'.format(processor)
        switchremoval.sheet.cell(row=14, column=10).value = \
            '{}'.format(chargermake)
        switchremoval.sheet.cell(row=15, column=10).value = \
            '{}'.format(chargermodel)
        switchremoval.sheet.cell(row=14, column=29).value = \
            '{}'.format(chargervolt)
        switchremoval.sheet.cell(row=15, column=29).value = \
            '{}'.format(chargeramp)
        switchremoval.sheet.cell(row=16, column=9).value = \
            '{}'.format(battmake)
        switchremoval.sheet.cell(row=17, column=9).value = \
            '{}'.format(battmodel)
        switchremoval.sheet.cell(row=18, column=9).value = \
            '{}'.format(battnum)
        switchremoval.sheet.cell(row=16, column=29).value = \
            '{}'.format(battvolt)
        switchremoval.sheet.cell(row=17, column=26).value = \
            '{}'.format(batttype)
        switchremoval.sheet.cell(row=18, column=19).value = \
            '{}'.format(battamp)
        switchremoval.sheet.cell(row=25, column=1).value = \
            '{}'.format(comments)

        # SET DIRECTORY

        directory = os.path.abspath('/sdcard/UpDown')  # FOR ANDROID VERSION

        # directory = os.path.abspath("/home/cbooth/UpDown") # FOR LINUX VERSION

        if os.path.exists(directory):
            rtuinstall.wb.save('{}/RTU {} - {}.xlsx'.format(directory,
                               serialnum, change))
            file_from = '{}/RTU {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = '/Switch/RTU {} - {} - {}.xlsx'.format(serialnum,
                    equipnum, change)
        else:
            os.makedirs(directory)
            rtuinstall.wb.save('{}/RTU {} - {}.xlsx'.format(directory,
                               serialnum, change))
            file_from = '{}/RTU {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/RTU {} - {} - {}.xlsx'.format(serialnum,
                    equipnum, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.direction.text = ''
        self.ids.streetpos.text = ''
        self.ids.locstatus.text = ''


class SwitchRemoval(Screen):

    def __init__(self, **kwargs):
        super(SwitchRemoval, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleremoval.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleremoval.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        polenum = self.ids.polenum.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        locname = self.ids.locname.text
        owner = self.ids.owner.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        streetpos = self.ids.streetpos.text
        locstatus = self.ids.locstatus.text
        equipnum = self.ids.equipnum.text
        serialnum = self.ids.serialnum.text
        manufacturer = self.ids.manufacturer.text
        stock = self.ids.stock.text
        status = self.ids.status.text
        switchtype = self.ids.switchtype.text
        equipreturn = self.ids.equipreturn.text

        # Push to new excel file

        switchremoval.sheet.cell(row=2, column=5).value = \
            '{}'.format(PMO)
        switchremoval.sheet.cell(row=2, column=14).value = \
            '{}'.format(change)
        switchremoval.sheet.cell(row=2, column=25).value = \
            '{}'.format('ENTERA')
        switchremoval.sheet.cell(row=4, column=4).value = \
            '{}'.format(polenum)
        switchremoval.sheet.cell(row=3, column=5).value = \
            '{}'.format(funcnum)
        switchremoval.sheet.cell(row=3, column=21).value = \
            '{}'.format(locnum)
        switchremoval.sheet.cell(row=4, column=21).value = \
            '{}'.format(locname)
        switchremoval.sheet.cell(row=3, column=33).value = \
            '{}'.format(owner)
        switchremoval.sheet.cell(row=5, column=4).value = \
            '{}'.format(streetnum)
        switchremoval.sheet.cell(row=5, column=11).value = \
            '{}'.format(streetname)
        switchremoval.sheet.cell(row=6, column=6).value = \
            '{}'.format(direction)
        switchremoval.sheet.cell(row=7, column=6).value = \
            '{}'.format(streetpos)
        switchremoval.sheet.cell(row=9, column=6).value = \
            '{}'.format(locstatus)
        switchremoval.sheet.cell(row=8, column=6).value = \
            '{}'.format(loctype)

        switchremoval.sheet.cell(row=20, column=5).value = \
            '{}'.format(equipnum)
        switchremoval.sheet.cell(row=20, column=23).value = \
            '{}'.format(serialnum)
        switchremoval.sheet.cell(row=21, column=5).value = \
            '{}'.format(manufacturer)
        switchremoval.sheet.cell(row=21, column=24).value = \
            '{}'.format(stock)
        switchremoval.sheet.cell(row=22, column=10).value = \
            '{}'.format(status)
        switchremoval.sheet.cell(row=23, column=5).value = \
            '{}'.format(switchtype)
        switchremoval.sheet.cell(row=24, column=5).value = \
            '{}'.format(equipreturn)

        # SET DIRECTORY

        directory = os.path.abspath('/sdcard/UpDown')  # FOR ANDROID VERSION

        # directory = os.path.abspath("/home/cbooth/UpDown") # FOR LINUX VERSION

        if os.path.exists(directory):
            switchremoval.wb.save('{}/Switch {} - {}.xlsx'.format(directory,
                                  serialnum, change))
            file_from = '{}/Switch {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/Switch {} - {} - {}.xlsx'.format(serialnum,
                    status, change)
        else:
            os.makedirs(directory)
            switchremoval.wb.save('{}/Switch {} - {}.xlsx'.format(directory,
                                  serialnum, change))
            file_from = '{}/Switch {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/Switch {} - {} - {}.xlsx'.format(serialnum,
                    status, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.direction.text = ''
        self.ids.streetpos.text = ''
        self.ids.locstatus.text = ''


class SwitchInstall(Screen):

    def __init__(self, **kwargs):
        super(SwitchInstall, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleremoval.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleremoval.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        polenum = self.ids.polenum.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        locname = self.ids.locname.text
        owner = self.ids.owner.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        streetpos = self.ids.streetpos.text
        locstatus = self.ids.locstatus.text
        loctype = self.ids.loctype.text
        equipnum = self.ids.equipnum.text
        serialnum = self.ids.serialnum.text
        manufacturer = self.ids.manufacturer.text
        stock = self.ids.stock.text
        ratedvolt = self.ids.ratedvolt.text
        ratedamp = self.ids.ratedamp.text
        interruptcurrent = self.ids.interruptcurrent.text
        phase = self.ids.phase.text
        switchingmedium = self.ids.switchingmedium.text
        switchtype = self.ids.switchtype.text
        addtype = self.ids.addtype.text
        motor = self.ids.motor.text
        scadamate = self.ids.scadamate.text
        conscada = self.ids.conscada.text
        status = self.ids.status.text

        # Push to new excel file

        switchinstall.sheet.cell(row=2, column=5).value = \
            '{}'.format(PMO)
        switchinstall.sheet.cell(row=2, column=14).value = \
            '{}'.format(change)
        switchinstall.sheet.cell(row=2, column=25).value = \
            '{}'.format('ENTERA')
        switchinstall.sheet.cell(row=4, column=4).value = \
            '{}'.format(polenum)
        switchinstall.sheet.cell(row=3, column=5).value = \
            '{}'.format(funcnum)
        switchinstall.sheet.cell(row=3, column=21).value = \
            '{}'.format(locnum)
        switchinstall.sheet.cell(row=4, column=21).value = \
            '{}'.format(locname)
        switchinstall.sheet.cell(row=3, column=33).value = \
            '{}'.format(owner)
        switchinstall.sheet.cell(row=5, column=4).value = \
            '{}'.format(streetnum)
        switchinstall.sheet.cell(row=5, column=11).value = \
            '{}'.format(streetname)
        switchinstall.sheet.cell(row=6, column=6).value = \
            '{}'.format(direction)
        switchinstall.sheet.cell(row=7, column=6).value = \
            '{}'.format(streetpos)
        switchinstall.sheet.cell(row=9, column=6).value = \
            '{}'.format(locstatus)
        switchinstall.sheet.cell(row=8, column=6).value = \
            '{}'.format(loctype)

        switchinstall.sheet.cell(row=11, column=5).value = \
            '{}'.format(equipnum)
        switchinstall.sheet.cell(row=11, column=23).value = \
            '{}'.format(serialnum)
        switchinstall.sheet.cell(row=12, column=5).value = \
            '{}'.format(manufacturer)
        switchinstall.sheet.cell(row=12, column=24).value = \
            '{}'.format(stock)
        switchinstall.sheet.cell(row=13, column=7).value = \
            '{}'.format(ratedvolt)
        switchinstall.sheet.cell(row=13, column=19).value = \
            '{}'.format(ratedamp)
        switchinstall.sheet.cell(row=13, column=31).value = \
            '{}'.format(interruptcurrent)
        switchinstall.sheet.cell(row=14, column=4).value = \
            '{}'.format(phase)
        switchinstall.sheet.cell(row=14, column=20).value = \
            '{}'.format(switchingmedium)
        switchinstall.sheet.cell(row=15, column=5).value = \
            '{}'.format(switchtype)
        switchinstall.sheet.cell(row=16, column=8).value = \
            '{}'.format(addtype)
        switchinstall.sheet.cell(row=17, column=5).value = \
            '{}'.format(motor)
        switchinstall.sheet.cell(row=17, column=7).value = \
            '{}'.format(scadamate)
        switchinstall.sheet.cell(row=17, column=31).value = \
            '{}'.format(conscada)
        switchinstall.sheet.cell(row=18, column=10).value = \
            '{}'.format(status)

        # SET DIRECTORY

        directory = os.path.abspath('/sdcard/UpDown')  # FOR ANDROID VERSION

        # directory = os.path.abspath("/home/cbooth/UpDown") # FOR LINUX VERSION

        if os.path.exists(directory):
            switchinstall.wb.save('{}/Switch {} - {}.xlsx'.format(directory,
                                  serialnum, change))
            file_from = '{}/Switch {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/Switch {} - {} - {}.xlsx'.format(serialnum,
                    status, change)
        else:
            os.makedirs(directory)
            switchinstall.wb.save('{}/Switch {} - {}.xlsx'.format(directory,
                                  serialnum, change))
            file_from = '{}/Switch {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/Switch {} - {} - {}.xlsx'.format(serialnum,
                    status, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.direction.text = ''
        self.ids.streetpos.text = ''
        self.ids.locstatus.text = ''


class NetTransRemoval(Screen):

    def __init__(self, **kwargs):
        super(NetTransRemoval, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleremoval.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleremoval.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        funcnum = self.ids.funcnum.text
        locname = self.ids.locname.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        locstatus = self.ids.locstatus.text
        equipnum = self.ids.equipnum.text
        serialnum = self.ids.serialnum.text
        stock = self.ids.stock.text
        manufacturer = self.ids.manufacturer.text
        kva = self.ids.kva.text
        year = self.ids.year.text
        feeder = self.ids.feeder.text
        status = self.ids.status.text
        equipreturn = self.ids.equipreturn.text

        # Push to new excel file

        nettransremoval.sheet.cell(row=2, column=5).value = \
            '{}'.format(PMO)
        nettransremoval.sheet.cell(row=2, column=17).value = \
            '{}'.format(change)
        nettransremoval.sheet.cell(row=2, column=29).value = \
            '{}'.format('ENTERA')
        nettransremoval.sheet.cell(row=3, column=5).value = \
            '{}'.format(funcnum)
        nettransremoval.sheet.cell(row=3, column=23).value = \
            '{}'.format(locnum)
        nettransremoval.sheet.cell(row=4, column=4).value = \
            '{}'.format(streetnum)
        nettransremoval.sheet.cell(row=4, column=9).value = \
            '{}'.format(streetname)
        nettransremoval.sheet.cell(row=4, column=27).value = \
            '{}'.format(direction)
        nettransremoval.sheet.cell(row=5, column=6).value = \
            '{}'.format(locstatus)
        nettransremoval.sheet.cell(row=13, column=5).value = \
            '{}'.format(equipnum)
        nettransremoval.sheet.cell(row=13, column=17).value = \
            '{}'.format(serialnum)
        nettransremoval.sheet.cell(row=13, column=31).value = \
            '{}'.format(stock)
        nettransremoval.sheet.cell(row=14, column=5).value = \
            '{}'.format(manufacturer)
        nettransremoval.sheet.cell(row=14, column=19).value = \
            '{}'.format(kva)
        nettransremoval.sheet.cell(row=14, column=27).value = \
            '{}'.format(year)
        nettransremoval.sheet.cell(row=14, column=34).value = \
            '{}'.format(feeder)
        nettransremoval.sheet.cell(row=15, column=7).value = \
            '{}'.format(status)
        nettransremoval.sheet.cell(row=15, column=24).value = \
            '{}'.format(equipreturn)

        # SET DIRECTORY

        directory = os.path.abspath('/sdcard/UpDown')  # FOR ANDROID VERSION

        # directory = os.path.abspath("/home/cbooth/UpDown") # FOR LINUX VERSION

        if os.path.exists(directory):
            nettransremoval.wb.save('{}/Network Transformer {} - {}.xlsx'.format(directory,
                                    serialnum, change))
            file_from = \
                '{}/Network Transformer {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/Network Transformer {} - {} - {}.xlsx'.format(serialnum,
                    status, change)
        else:
            os.makedirs(directory)
            nettransremoval.wb.save('{}/Network Transformer {} - {}.xlsx'.format(directory,
                                    serialnum, change))
            file_from = \
                '{}/Network Transformer {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/Network Transformer {} - {} - {}.xlsx'.format(serialnum,
                    status, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.direction.text = ''
        self.ids.streetpos.text = ''
        self.ids.locstatus.text = ''


class NetTransInstall(Screen):

    def __init__(self, **kwargs):
        super(NetTransInstall, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleremoval.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleremoval.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        locname = self.ids.locname.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        locstatus = self.ids.locstatus.text
        equipnum = self.ids.equipnum.text
        serialnum = self.ids.serialnum.text
        stock = self.ids.stock.text
        manufacturer = self.ids.manufacturer.text
        kva = self.ids.kva.text
        year = self.ids.year.text
        feeder = self.ids.feeder.text
        secondvolt = self.ids.secondvolt.text
        status = self.ids.status.text

        # Push to new excel file

        nettransinstall.sheet.cell(row=2, column=5).value = \
            '{}'.format(PMO)
        nettransinstall.sheet.cell(row=2, column=17).value = \
            '{}'.format(change)
        nettransinstall.sheet.cell(row=2, column=29).value = \
            '{}'.format('ENTERA')
        nettransinstall.sheet.cell(row=3, column=5).value = \
            '{}'.format(funcnum)
        nettransinstall.sheet.cell(row=3, column=23).value = \
            '{}'.format(locnum)
        nettransinstall.sheet.cell(row=4, column=4).value = \
            '{}'.format(streetnum)
        nettransinstall.sheet.cell(row=4, column=9).value = \
            '{}'.format(streetname)
        nettransinstall.sheet.cell(row=4, column=27).value = \
            '{}'.format(direction)
        nettransinstall.sheet.cell(row=5, column=6).value = \
            '{}'.format(locstatus)
        nettransinstall.sheet.cell(row=2, column=5).value = \
            '{}'.format(equipnum)
        nettransinstall.sheet.cell(row=2, column=17).value = \
            '{}'.format(serialnum)
        nettransinstall.sheet.cell(row=2, column=29).value = \
            '{}'.format(stock)
        nettransinstall.sheet.cell(row=3, column=5).value = \
            '{}'.format(manufacturer)
        nettransinstall.sheet.cell(row=3, column=23).value = \
            '{}'.format(kva)
        nettransinstall.sheet.cell(row=4, column=4).value = \
            '{}'.format(year)
        nettransinstall.sheet.cell(row=4, column=9).value = \
            '{}'.format(feeder)
        nettransinstall.sheet.cell(row=4, column=27).value = \
            '{}'.format(secondvolt)
        nettransinstall.sheet.cell(row=5, column=6).value = \
            '{}'.format(status)

        # SET DIRECTORY

        directory = os.path.abspath('/sdcard/UpDown')  # FOR ANDROID VERSION

        # directory = os.path.abspath("/home/cbooth/UpDown") # FOR LINUX VERSION

        if os.path.exists(directory):
            nettransinstall.wb.save('{}/Network Transformer {} - {}.xlsx'.format(directory,
                                    serialnum, change))
            file_from = \
                '{}/Network Transformer {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/Network Transformer {} - {} - {}.xlsx'.format(serialnum,
                    status, change)
        else:
            os.makedirs(directory)
            nettransinstall.wb.save('{}/Network Transformer {} - {}.xlsx'.format(directory,
                                    serialnum, change))
            file_from = \
                '{}/Network Transformer {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/Network Transformer {} - {} - {}.xlsx'.format(serialnum,
                    status, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.direction.text = ''
        self.ids.streetpos.text = ''
        self.ids.locstatus.text = ''


class NetProInstall(Screen):

    def __init__(self, **kwargs):
        super(NetProInstall, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleremoval.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleremoval.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        locname = self.ids.locname.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        locstatus = self.ids.locstatus.text
        equipnum = self.ids.equipnum.text
        serialnum = self.ids.serialnum.text
        manufacturer = self.ids.manufacturer.text
        stock = self.ids.stock.text
        year = self.ids.year.text
        enclosure = self.ids.enclosure.text
        internalserial = self.ids.internalserial.text
        internalmanufacturer = self.ids.internalmanufacturer.text
        internalyear = self.ids.internalyear.text
        ctratio = self.ids.ctratio.text
        ratedvolt = self.ids.ratedvolt.text
        ratedcurrent = self.ids.ratedcurrent.text
        breaker = self.ids.breaker.text
        interruptcapacity = self.ids.interruptcapacity.text
        replacedbreakeronly = self.ids.replacedbreakeronly.text
        feeder = self.ids.feeder.text
        status = self.ids.status.text

        # Push to new excel file

        netproinstall.sheet.cell(row=2, column=5).value = \
            '{}'.format(PMO)
        netproinstall.sheet.cell(row=2, column=17).value = \
            '{}'.format(change)
        netproinstall.sheet.cell(row=2, column=29).value = \
            '{}'.format('ENTERA')
        netproinstall.sheet.cell(row=3, column=5).value = \
            '{}'.format(funcnum)
        netproinstall.sheet.cell(row=3, column=23).value = \
            '{}'.format(locnum)
        netproinstall.sheet.cell(row=4, column=4).value = \
            '{}'.format(streetnum)
        netproinstall.sheet.cell(row=4, column=9).value = \
            '{}'.format(streetname)
        netproinstall.sheet.cell(row=4, column=27).value = \
            '{}'.format(direction)
        netproinstall.sheet.cell(row=5, column=6).value = \
            '{}'.format(locstatus)
        netproinstall.sheet.cell(row=21, column=10).value = \
            '{}'.format(equipnum)
        netproinstall.sheet.cell(row=21, column=24).value = \
            '{}'.format(serialnum)
        netproinstall.sheet.cell(row=22, column=10).value = \
            '{}'.format(manufacturer)
        netproinstall.sheet.cell(row=22, column=23).value = \
            '{}'.format(stock)
        netproinstall.sheet.cell(row=22, column=34).value = \
            '{}'.format(year)
        netproinstall.sheet.cell(row=23, column=9).value = \
            '{}'.format(enclosure)
        netproinstall.sheet.cell(row=24, column=9).value = \
            '{}'.format(internalserial)
        netproinstall.sheet.cell(row=24, column=22).value = \
            '{}'.format(internalmanufacturer)
        netproinstall.sheet.cell(row=24, column=34).value = \
            '{}'.format(internalyear)
        netproinstall.sheet.cell(row=27, column=5).value = \
            '{}'.format(ctratio)
        netproinstall.sheet.cell(row=25, column=5).value = \
            '{}'.format(ratedvolt)
        netproinstall.sheet.cell(row=26, column=5).value = \
            '{}'.format(ratedcurrent)
        netproinstall.sheet.cell(row=25, column=31).value = \
            '{}'.format(breaker)
        netproinstall.sheet.cell(row=27, column=31).value = \
            '{}'.format(interruptcapacity)
        netproinstall.sheet.cell(row=28, column=5).value = \
            '{}'.format(replacedbreakeronly)
        netproinstall.sheet.cell(row=28, column=15).value = \
            '{}'.format(feeder)
        netproinstall.sheet.cell(row=28, column=28).value = \
            '{}'.format(status)

        # SET DIRECTORY

        directory = os.path.abspath('/sdcard/UpDown')  # FOR ANDROID VERSION

        # directory = os.path.abspath("/home/cbooth/UpDown") # FOR LINUX VERSION

        if os.path.exists(directory):
            netproinstall.wb.save('{}/Network Protector {} - {}.xlsx'.format(directory,
                                  serialnum, change))
            file_from = \
                '{}/Network Protector {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/Network Protector {} - {} - {}.xlsx'.format(serialnum,
                    status, change)
        else:
            os.makedirs(directory)
            netproinstall.wb.save('{}/Network Protector {} - {}.xlsx'.format(directory,
                                  polenum, change))
            file_from = \
                '{}/Network Protector {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/Network Protector {} - {} - {}.xlsx'.format(serialnum,
                    status, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.direction.text = ''
        self.ids.locstatus.text = ''


class NetProRemoval(Screen):

    def __init__(self, **kwargs):
        super(NetProRemoval, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def show_load2(self):
        content = LoadDialog2(load2=self.load2,
                              cancel=self.dismiss_popup)
        self._popup = Popup(title='Load file', content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        netproinstall.bsheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def load2(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        netproinstall.asheet.add_image(img)
        stream.flush()
        self.dismiss_popup()

    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        locname = self.ids.locname.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        locstatus = self.ids.locstatus.text
        equipnum = self.ids.equipnum.text
        serialnum = self.ids.serialnum.text
        manufacturer = self.ids.manufacturer.text
        stock = self.ids.stock.text
        year = self.ids.year.text
        feeder = self.ids.feeder.text
        breaker = self.ids.breaker.text
        status = self.ids.status.text
        equipreturn = self.ids.equipreturn.text

        # Push to new excel file

        netproremoval.sheet.cell(row=2, column=5).value = \
            '{}'.format(PMO)
        netproremoval.sheet.cell(row=2, column=17).value = \
            '{}'.format(change)
        netproremoval.sheet.cell(row=2, column=29).value = \
            '{}'.format('ENTERA')
        netproremoval.sheet.cell(row=3, column=5).value = \
            '{}'.format(funcnum)
        netproremoval.sheet.cell(row=3, column=23).value = \
            '{}'.format(locnum)
        netproremoval.sheet.cell(row=4, column=4).value = \
            '{}'.format(streetnum)
        netproremoval.sheet.cell(row=4, column=9).value = \
            '{}'.format(streetname)
        netproremoval.sheet.cell(row=4, column=27).value = \
            '{}'.format(direction)
        netproremoval.sheet.cell(row=5, column=6).value = \
            '{}'.format(locstatus)
        netproremoval.sheet.cell(row=30, column=5).value = \
            '{}'.format(equipnum)
        netproremoval.sheet.cell(row=30, column=23).value = \
            '{}'.format(serialnum)
        netproremoval.sheet.cell(row=31, column=5).value = \
            '{}'.format(manufacturer)
        netproremoval.sheet.cell(row=31, column=24).value = \
            '{}'.format(stock)
        netproremoval.sheet.cell(row=32, column=3).value = \
            '{}'.format(year)
        netproremoval.sheet.cell(row=32, column=13).value = \
            '{}'.format(feeder)
        netproremoval.sheet.cell(row=32, column=23).value = \
            '{}'.format(breaker)
        netproremoval.sheet.cell(row=33, column=7).value = \
            '{}'.format(status)
        netproremoval.sheet.cell(row=33, column=24).value = \
            '{}'.format(equipreturn)

        # SET DIRECTORY
        # directory = os.path.abspath("/sdcard/UpDown") #FOR ANDROID VERSION

        directory = os.path.abspath('/home/cbooth/UpDown')  # FOR LINUX VERSION
        if os.path.exists(directory):
            netproremoval.wb.save('{}/Protector {} - {}.xlsx'.format(directory,
                                  serialnum, change))
            file_from = '{}/Protector {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/Protector {} - {} - {}.xlsx'.format(serialnum,
                    status, change)
        else:
            os.makedirs(directory)
            netproremoval.wb.save('{}/Protector {} - {}.xlsx'.format(directory,
                                  serialnum, change))
            file_from = '{}/Protector {} - {}.xlsx'.format(directory,
                    serialnum, change)
            file_to = \
                '/SOFTWARE/Protector {} - {} - {}.xlsx'.format(serialnum,
                    status, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''


sm = ScreenManager(transition=NoTransition())  # Creates an instance (copy) of ScreenManager as variable sm. ScreenManager switches between Screen Objects.
sm.add_widget(mainscreen.WelcomeScreen(name='welcome_screen'))  # Adds WelcomeScreen widget to ScreenManager. ScreenManager id's screen as welcome_screen.
sm.add_widget(PoleRemoval(name='poleremove'))
sm.add_widget(PoleInstall(name='poleinstall'))
sm.add_widget(StreetlightRemoval(name='streetlightremoval'))
sm.add_widget(StreetlightInstall(name='streetlightinstall'))
sm.add_widget(TransformerRemoval(name='transformerremoval'))
sm.add_widget(TransformerInstall(name='transformerinstall'))
sm.add_widget(RTURemoval(name='rturemoval'))
sm.add_widget(RTUInstall(name='rtuinstall'))
sm.add_widget(SwitchRemoval(name='switchremoval'))
sm.add_widget(SwitchInstall(name='switchinstall'))
sm.add_widget(NetTransRemoval(name='nettransremoval'))
sm.add_widget(NetTransInstall(name='nettransinstall'))
sm.add_widget(NetProRemoval(name='netproremoval'))
sm.add_widget(NetProInstall(name='netproinstall'))


class SwitchingScreenApp(App):  # Creates the instance (copy) of the Kivy App class named SwitchingScreenApp

    def build(self):  # build is a method of Kivy's App class used to place widgets onto the GUI.
        return sm  # return calls the build method which in turn builds the GUI.'


SwitchingScreenApp().run()  # Runs SwitchingScreenApp
