#!/usr/bin/python
# -*- coding: utf-8 -*-
import kivy  # Required to run Kivy such as the next line of code
kivy.require('1.9.1')  # used to alert user if this code is run on an earlier version of Kivy.
from kivy.app import App  # Imports the base App class required for Kivy Apps
from kivy.lang import Builder  # Imports the KV language builder that provides the layout of kivy screens
from kivy.uix.screenmanager import ScreenManager, Screen, NoTransition  # Imports the Kivy Screen manager and Kivys Screen class
from kivy.core.window import Window
import mainscreen
import poleremoval
import poleinstall
import os
import dropbox
from openpyxl import load_workbook, drawing
import openpyxl
from kivy.properties import ObjectProperty
from kivy.uix.popup import Popup
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.image import Image
from kivy.uix.boxlayout import BoxLayout
import datepicker
#import filefile
import loadload
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.scrollview import ScrollView
from textwrap import dedent
from kivy.uix.button import Button
from kivy.properties import ListProperty

from plyer import filechooser

TOKEN = \
    'S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
dbx = dropbox.Dropbox(
    TOKEN,
    max_retries_on_error=10,
    max_retries_on_rate_limit=None,
    user_agent=None,
    session=None,
    headers=None,
    timeout=10000,
    )
Window.softinput_mode = 'below_target'
#Builder.load_file('switchingscreen.kv')


class LoadDialog(FloatLayout):

    load = ObjectProperty(None)
    cancel = ObjectProperty(None)


class LoadDialog2(FloatLayout):

    toad = ObjectProperty(None)
    cancel = ObjectProperty(None)


class PoleRemoval(Screen):
	

    def __init__(self, **kwargs):
        super(PoleRemoval, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title='', content=content,
                            size_hint=(0.5, 0.5))
        self._popup.open()


    def load(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        #poleremoval.wb.create_sheet('Photo').add_image(img)
        stream.flush()



    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        entera = self.ids.entera.text
        PMO = self.ids.PMO.text
        change = self.ids.change.text
        polenum = self.ids.polenum.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        locname = self.ids.locname.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        streetpos = self.ids.streetpos.text
        locstatus = self.ids.locstatus.text
        poleheightremoval = self.ids.poleheightremoval.text
        polematerial = self.ids.polematerial.text
        stub = self.ids.stub.text
        comments = self.ids.comments.text

        # Push to new excel file

        poleremoval.sheet.cell(row=2, column=2).value = '{}'.format(PMO)
        poleremoval.sheet.cell(row=2, column=6).value = \
            '{}'.format(change)
        poleremoval.sheet.cell(row=2, column=9).value = \
            '{}'.format('ENTERA')
        poleremoval.sheet.cell(row=3, column=2).value = \
            '{}'.format(polenum)
        poleremoval.sheet.cell(row=3, column=5).value = \
            '{}'.format(funcnum)
        poleremoval.sheet.cell(row=3, column=9).value = \
            '{}'.format(locnum)
        poleremoval.sheet.cell(row=4, column=2).value = \
            '{}'.format(locname)
        poleremoval.sheet.cell(row=5, column=2).value = \
            '{}'.format(streetnum)
        poleremoval.sheet.cell(row=5, column=5).value = \
            '{}'.format(streetname)
        poleremoval.sheet.cell(row=6, column=2).value = \
            '{}'.format(direction)
        poleremoval.sheet.cell(row=7, column=3).value = \
            '{}'.format(streetpos)
        poleremoval.sheet.cell(row=8, column=3).value = \
            '{}'.format(locstatus)
        poleremoval.sheet.cell(row=58, column=3).value = \
            '{}'.format(poleheightremoval)
        poleremoval.sheet.cell(row=58, column=6).value = \
            '{}'.format(polematerial)
        poleremoval.sheet.cell(row=58, column=10).value = \
            '{}'.format(stub)
        poleremoval.sheet.cell(row=60, column=1).value = \
            '{}'.format(comments)

        directory = os.path.abspath('/sdcard/UpDown')

        #directory = os.path.abspath("/home/cbooth/UpDown")

        if os.path.exists(directory):
            poleremoval.wb.save('{}/{} - Pole {} - {}.xlsx'.format(directory,
                                entera, polenum, change))
            file_from = '{}/{} - Pole {} - {}.xlsx'.format(directory,
                    entera, polenum, change)
            file_to = \
                '/CHANGEOUT/{} - Pole {} - {} - {}.xlsx'.format(entera,
                    polenum, streetname, change)
        else:
            os.makedirs(directory)
            poleremoval.wb.save('{}/{} - Pole {} - {}.xlsx'.format(directory,
                                entera, polenum, change))
            file_from = '{}/{} - Pole {} - {}.xlsx'.format(directory,
                    entera, polenum, change)
            file_to = \
                '/CHANGEOUT/{} - Pole {} - {} - {}.xlsx'.format(entera,
                    polenum, streetname, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)

        # file_from = ''

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''

        self.ids.directionspinner.text = ''

        self.ids.streetpos.text = ''
        self.ids.locstatus.text = ''
        self.ids.poleheightremoval.text = ''
        self.ids.polematerial.text = ''
        self.ids.stub.text = ''
        self.ids.comments.text = ''

    loadfile = ObjectProperty(None)
    savefile = ObjectProperty(None)
    text_input = ObjectProperty(None)


class PoleInstall(Screen):  # Creates a SecondScreen widget from the above kv language data string.

    def __init__(self, **kwargs):
        super(PoleInstall, self).__init__(**kwargs)

    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        content = LoadDialog2(load=self.toad, cancel=self.dismiss_popup)
        self._popup = Popup(title='', content=content,
                            size_hint=(0.5, 0.5))
        self._popup.open()
        
    def show_load2(self):
        content = LoadDialog2(toad=self.toad, cancel=self.dismiss_popup)
        self._popup = Popup(title='', content=content,
                            size_hint=(0.5, 0.5))
        self._popup.open()

    def toad(self, path, filename):
        stream = open(os.path.join(path, filename[0]), 'rb')
        img = openpyxl.drawing.image.Image(stream)
        img.anchor = 'A1'
        poleinstall.wb.create_sheet('Photo').add_image(img)
        stream.flush()




    def on_spinner_select(self, text):
        direction = self.ids.directionspinner.text

    def pressed(self):
        dbx = \
            dropbox.Dropbox('S0orUPN1C3AAAAAAAAEeutG2Re5VQr7CcNFDKBvW0CbeTRCNpCRHIlLTbW2De7DH'
                            )

      # Get the value of all of the text inputs

        PMO = self.ids.PMO.text
        change = self.ids.change.text
        polenum = self.ids.polenum.text
        funcnum = self.ids.funcnum.text
        locnum = self.ids.locnum.text
        locname = self.ids.locname.text
        streetnum = self.ids.streetnum.text
        streetname = self.ids.streetname.text
        direction = self.ids.directionspinner.text
        streetpos = self.ids.streetpos.text
        locstatus = self.ids.locstatus.text
        poleowner = self.ids.poleowner.text
        poleheightinstall = self.ids.poleheightinstall.text
        stocknum = self.ids.stocknum.text
        materialinstall = self.ids.polematerial.text
        polesetting = self.ids.polesetting.text
        poleframing = self.ids.poleframing.text
        surface = self.ids.surface.text
        jointuse = self.ids.jointuse.text
        traffic = self.ids.traffic.text
        topext = self.ids.topext.text
        feeder = self.ids.feeder.text
        streetlight = self.ids.streetlight.text
        riser = self.ids.riser.text
        suspension = self.ids.suspension.text
        signs = self.ids.signs.text
        guy = self.ids.guy.text

        # Push to new excel file

        poleinstall.sheet.cell(row=2, column=2).value = '{}'.format(PMO)
        poleinstall.sheet.cell(row=2, column=6).value = \
            '{}'.format(change)
        poleinstall.sheet.cell(row=2, column=9).value = \
            '{}'.format('ENTERA')
        poleinstall.sheet.cell(row=3, column=2).value = \
            '{}'.format(polenum)
        poleinstall.sheet.cell(row=3, column=5).value = \
            '{}'.format(funcnum)
        poleinstall.sheet.cell(row=3, column=9).value = \
            '{}'.format(locnum)
        poleinstall.sheet.cell(row=4, column=2).value = \
            '{}'.format(locname)
        poleinstall.sheet.cell(row=5, column=2).value = \
            '{}'.format(streetnum)
        poleinstall.sheet.cell(row=5, column=5).value = \
            '{}'.format(streetname)
        poleinstall.sheet.cell(row=6, column=2).value = \
            '{}'.format(direction)
        poleinstall.sheet.cell(row=7, column=3).value = \
            '{}'.format(streetpos)
        poleinstall.sheet.cell(row=8, column=3).value = \
            '{}'.format(locstatus)
        poleinstall.sheet.cell(row=10, column=3).value = \
            '{}'.format(poleowner)
        poleinstall.sheet.cell(row=11, column=3).value = \
            '{}'.format(poleheightinstall)
        poleinstall.sheet.cell(row=11, column=7).value = \
            '{}'.format(stocknum)
        poleinstall.sheet.cell(row=12, column=3).value = \
            '{}'.format(materialinstall)
        poleinstall.sheet.cell(row=13, column=3).value = \
            '{}'.format(polesetting)
        poleinstall.sheet.cell(row=14, column=3).value = \
            '{}'.format(poleframing)
        poleinstall.sheet.cell(row=15, column=3).value = \
            '{}'.format(surface)
        poleinstall.sheet.cell(row=15, column=7).value = \
            '{}'.format(jointuse)
        poleinstall.sheet.cell(row=15, column=9).value = \
            '{}'.format(traffic)
        poleinstall.sheet.cell(row=16, column=3).value = \
            '{}'.format(topext)
        poleinstall.sheet.cell(row=16, column=7).value = \
            '{}'.format(feeder)
        poleinstall.sheet.cell(row=16, column=9).value = \
            '{}'.format(streetlight)
        poleinstall.sheet.cell(row=17, column=3).value = \
            '{}'.format(riser)
        poleinstall.sheet.cell(row=17, column=7).value = \
            '{}'.format(suspension)
        poleinstall.sheet.cell(row=17, column=9).value = \
            '{}'.format(signs)
        poleinstall.sheet.cell(row=18, column=3).value = \
            '{}'.format(guy)

        # SET DIRECTORY

        directory = os.path.abspath('/sdcard/UpDown')  # FOR ANDROID VERSION

        #directory = os.path.abspath("/home/cbooth/UpDown") # FOR LINUX VERSION

        if os.path.exists(directory):
            poleinstall.wb.save('{}/Pole {} - {}.xlsx'.format(directory,
                                polenum, change))
            file_from = '{}/Pole {} - {}.xlsx'.format(directory,
                    polenum, change)
            file_to = \
                '/CHANGEOUT/Pole {} - {} - {}.xlsx'.format(polenum,
                    streetname, change)
        else:
            os.makedirs(directory)
            poleinstall.wb.save('{}/Pole {} - {}.xlsx'.format(directory,
                                polenum, change))
            file_from = '{}/Pole {} - {} - {}.xlsx'.format(directory,
                    polenum, streetname, change)
            file_to = \
                '/CHANGEOUT/Pole {} - {} - {}.xlsx'.format(polenum,
                    streetname, change)

        with open(file_from, 'rb') as f:
            dbx.files_alpha_upload(f.read(), file_to)  # Open the file as a binary stream and push to dropbox

        # Reset fields to no value

        self.ids.PMO.text = ''
        self.ids.change.text = ''
        self.ids.polenum.text = ''
        self.ids.funcnum.text = ''
        self.ids.locnum.text = ''
        self.ids.locname.text = ''
        self.ids.streetnum.text = ''
        self.ids.streetname.text = ''
        self.ids.directionspinner.text = ''
        self.ids.streetpos.text = ''
        self.ids.locstatus.text = ''


class FileChoose(Button):
    '''
    Button that triggers 'filechooser.open_file()' and processes
    the data response from filechooser Activity.
    '''

    selection = ListProperty([])

    def choose(self):
        '''
        Call plyer filechooser API to run a filechooser Activity.
        '''
        filechooser.open_file(on_selection=self.handle_selection)

    def handle_selection(self, selection):
        '''
        Callback function for handling the selection response from Activity.
        '''
        loadload.load (self, selection)

    def on_selection(self, *a, **k):
        '''
        Update TextInput.text after FileChoose.selection is changed
        via FileChoose.handle_selection.
        '''
        App.get_running_app().root.ids.result.text = str(self.selection)
        
        ##################
        
    def choose2(self):
        '''
        Call plyer filechooser API to run a filechooser Activity.
        '''
        filechooser.open_file(on_selection=self.handle_selection2)

    def handle_selection2(self, selection):
        '''
        Callback function for handling the selection response from Activity.
        '''
        loadload.toad (self, selection)

    def on_selection2(self, *a, **k):
        '''
        Update TextInput.text after FileChoose.selection is changed
        via FileChoose.handle_selection.
        '''
        App.get_running_app().root.ids.result.text = str(self.selection)
        
        ###############################################################
        


sm = ScreenManager(transition=NoTransition())  # Creates an instance (copy) of ScreenManager as variable sm. ScreenManager switches between Screen Objects.
sm.add_widget(mainscreen.WelcomeScreen(name='welcome_screen'))  # Adds WelcomeScreen widget to ScreenManager. ScreenManager id's screen as welcome_screen.
sm.add_widget(PoleRemoval(name='poleremove'))
sm.add_widget(PoleInstall(name='poleinstall'))


class SwitchingScreenApp(App):  # Creates the instance (copy) of the Kivy App class named SwitchingScreenApp

    def build(self):  # build is a method of Kivy's App class used to place widgets onto the GUI.
        sm = ScreenManager(transition=NoTransition())  # Creates an instance (copy) of ScreenManager as variable sm. ScreenManager switches between Screen Objects.
        sm.add_widget(mainscreen.WelcomeScreen(name='welcome_screen'))  # Adds WelcomeScreen widget to ScreenManager. ScreenManager id's screen as welcome_screen.
        sm.add_widget(PoleRemoval(name='poleremove'))
        sm.add_widget(PoleInstall(name='poleinstall'))
        return sm  # return calls the build method which in turn builds the GUI.'


SwitchingScreenApp().run()  # Runs SwitchingScreenApp
