import kivy
from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.anchorlayout import AnchorLayout
import kivy  # Required to run Kivy such as the next line of code
kivy.require('1.9.1')  # used to alert user if this code is run on an earlier version of Kivy.
from kivy.app import App  # Imports the base App class required for Kivy Apps
from kivy.lang import Builder  # Imports the KV language builder that provides the layout of kivy screens
from kivy.uix.screenmanager import ScreenManager, Screen # Imports the Kivy Screen manager and Kivys Screen class
from kivy.properties import ObjectProperty, StringProperty
import os
from openpyxl import load_workbook, drawing
import openpyxl
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.boxlayout import BoxLayout

wb = load_workbook('excel/pole.xlsx') 
sheet = wb['Sheet2']
#bsheet = wb.create_sheet('Before Photo')
#asheet = wb.create_sheet('After Photo')

class PoleRemoval(Screen):
    def __init__(self, **kwargs):
        super(PoleRemoval, self).__init__(**kwargs)

